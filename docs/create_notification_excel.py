# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 1. 알림 항목 전체 목록
notifications_data = [
    # 3.1.1 학기별 지도 계획
    ['PLAN_STUDENT_REGISTERED', '학생 학기별 지도 계획 등록', '학생이 학기별 지도 계획을 신규 등록', '학생', '담당 교수', '높음', '논문 지도 계획', '[학생명]님이 [학기명] 학기별 지도 계획을 등록했습니다.'],
    ['PLAN_STUDENT_UPDATED', '학생 학기별 지도 계획 수정', '학생이 승인 전 계획을 수정', '학생', '담당 교수', '보통', '논문 지도 계획', '[학생명]님이 [학기명] 학기별 지도 계획을 수정했습니다.'],
    ['PLAN_PROFESSOR_APPROVED', '학기별 지도 계획 승인', '교수가 학생의 계획을 승인', '교수', '해당 학생', '높음', '논문 지도 계획', '담당 교수님이 [학기명] 학기별 지도 계획을 승인했습니다.'],
    ['PLAN_PROFESSOR_REJECTED', '학기별 지도 계획 반려', '교수가 학생의 계획을 반려', '교수', '해당 학생', '높음', '논문 지도 계획', '담당 교수님이 [학기명] 학기별 지도 계획을 반려했습니다. 사유: [반려사유]'],
    ['PLAN_WEEK_ADDED', '주차 추가', '학생이 계획에 새로운 주차 추가', '학생', '담당 교수', '낮음', '논문 지도 계획', '[학생명]님이 [학기명] 계획에 주차를 추가했습니다.'],
    ['PLAN_WEEK_DELETED', '주차 삭제', '학생이 계획에서 주차 삭제', '학생', '담당 교수', '낮음', '논문 지도 계획', '[학생명]님이 [학기명] 계획에서 주차를 삭제했습니다.'],

    # 3.2.1 논문 지도 요청
    ['GUIDANCE_REQUEST_CREATED', '논문 지도 요청 등록', '학생이 새로운 논문 지도 요청 등록', '학생', '담당 교수', '높음', '논문 지도 활동', '[학생명]님이 논문 지도를 요청했습니다. 제목: [논문제목]'],
    ['GUIDANCE_REQUEST_UPDATED', '논문 지도 요청 수정', '학생이 피드백 전 지도 요청 수정', '학생', '담당 교수', '보통', '논문 지도 활동', '[학생명]님이 논문 지도 요청을 수정했습니다.'],
    ['GUIDANCE_FEEDBACK_CREATED', '논문 지도 피드백 등록', '교수가 학생의 지도 요청에 피드백 등록', '교수', '해당 학생', '높음', '논문 지도 활동', '담당 교수님이 논문 지도 피드백을 등록했습니다.'],
    ['GUIDANCE_FEEDBACK_UPDATED', '논문 지도 피드백 수정', '교수가 기존 피드백 수정', '교수', '해당 학생', '보통', '논문 지도 활동', '담당 교수님이 논문 지도 피드백을 수정했습니다.'],
    ['GUIDANCE_DOCUMENT_UPLOADED', '지도 자료 업로드', '학생이 논문 지도 요청에 파일 첨부', '학생', '담당 교수', '보통', '논문 지도 활동', '[학생명]님이 논문 지도 요청에 자료를 업로드했습니다.'],
    ['GUIDANCE_FEEDBACK_DOCUMENT_UPLOADED', '피드백 자료 업로드', '교수가 피드백에 파일 첨부', '교수', '해당 학생', '보통', '논문 지도 활동', '담당 교수님이 피드백에 자료를 업로드했습니다.'],

    # 3.2.2 실시간 지도 예약
    ['RESERVATION_CREATED', '실시간 지도 예약 등록', '학생이 실시간 지도 예약', '학생', '담당 교수', '높음', '논문 지도 활동', '[학생명]님이 [날짜] [시간]에 실시간 지도를 예약했습니다.'],
    ['RESERVATION_CANCELLED', '실시간 지도 예약 취소', '학생이 예약 취소', '학생', '담당 교수', '높음', '논문 지도 활동', '[학생명]님이 [날짜] [시간] 예약을 취소했습니다.'],
    ['RESERVATION_CONFIRMED', '실시간 지도 예약 승인', '교수가 예약 승인', '교수', '해당 학생', '높음', '논문 지도 활동', '담당 교수님이 [날짜] [시간] 예약을 승인했습니다.'],
    ['RESERVATION_REJECTED', '실시간 지도 예약 거절', '교수가 예약 거절', '교수', '해당 학생', '높음', '논문 지도 활동', '담당 교수님이 [날짜] [시간] 예약을 거절했습니다.'],
    ['RESERVATION_REMINDER', '실시간 지도 예약 알림', '예약 시간 1시간 전 자동 발송', '시스템', '학생, 교수', '높음', '논문 지도 활동', '1시간 후 [학생명]님과의 실시간 지도가 예정되어 있습니다.'],

    # 3.3.1 논문 신청
    ['THESIS_APPLICATION_SUBMITTED', '논문 신청 제출', '학생이 논문 신청', '학생', '담당 교수, 관리자', '높음', '논문 제출', '[학생명]님이 논문 신청을 제출했습니다.'],
    ['THESIS_APPLICATION_APPROVED', '논문 신청 승인', '관리자가 논문 신청 승인', '관리자', '해당 학생', '높음', '논문 제출', '논문 신청이 승인되었습니다.'],
    ['THESIS_APPLICATION_REJECTED', '논문 신청 반려', '관리자가 논문 신청 반려', '관리자', '해당 학생', '높음', '논문 제출', '논문 신청이 반려되었습니다. 사유: [반려사유]'],

    # 3.3.2 학위 논문 제출
    ['THESIS_SUBMITTED', '학위 논문 제출', '학생이 학위 논문 제출', '학생', '담당 교수, 관리자', '높음', '논문 제출', '[학생명]님이 학위 논문을 제출했습니다. 제목: [논문제목]'],
    ['THESIS_RESUBMITTED', '학위 논문 재제출', '학생이 반려 후 논문 재제출', '학생', '담당 교수, 관리자', '높음', '논문 제출', '[학생명]님이 학위 논문을 재제출했습니다.'],
    ['THESIS_APPROVED', '학위 논문 승인', '관리자/교수가 논문 승인', '관리자/교수', '해당 학생', '높음', '논문 제출', '학위 논문이 승인되었습니다.'],
    ['THESIS_REJECTED', '학위 논문 반려', '관리자/교수가 논문 반려', '관리자/교수', '해당 학생', '높음', '논문 제출', '학위 논문이 반려되었습니다. 사유: [반려사유]'],
    ['THESIS_DOCUMENT_UPDATED', '학위 논문 파일 수정', '학생이 제출 논문 파일 수정', '학생', '담당 교수, 관리자', '보통', '논문 제출', '[학생명]님이 학위 논문 파일을 수정했습니다.'],

    # 3.3.3 학술지 논문 제출
    ['JOURNAL_SUBMITTED', '학술지 논문 제출', '학생이 학술지 논문 제출', '학생', '담당 교수, 관리자', '높음', '논문 제출', '[학생명]님이 학술지 논문을 제출했습니다. 학술지명: [학술지명]'],
    ['JOURNAL_RESUBMITTED', '학술지 논문 재제출', '학생이 반려 후 논문 재제출', '학생', '담당 교수, 관리자', '높음', '논문 제출', '[학생명]님이 학술지 논문을 재제출했습니다.'],
    ['JOURNAL_APPROVED', '학술지 논문 승인', '관리자/교수가 논문 승인', '관리자/교수', '해당 학생', '높음', '논문 제출', '학술지 논문이 승인되었습니다.'],
    ['JOURNAL_REJECTED', '학술지 논문 반려', '관리자/교수가 논문 반려', '관리자/교수', '해당 학생', '높음', '논문 제출', '학술지 논문이 반려되었습니다. 사유: [반려사유]'],

    # 3.3.4 최종 논문 제목 등록
    ['TITLE_REGISTERED', '최종 논문 제목 등록', '학생이 최종 논문 제목 등록', '학생', '담당 교수, 관리자', '높음', '논문 제출', '[학생명]님이 최종 논문 제목을 등록했습니다. 제목: [논문제목]'],
    ['TITLE_UPDATED', '최종 논문 제목 수정', '학생이 논문 제목 수정', '학생', '담당 교수, 관리자', '보통', '논문 제출', '[학생명]님이 최종 논문 제목을 수정했습니다.'],
    ['TITLE_APPROVED', '최종 논문 제목 승인', '교수/관리자가 제목 승인', '교수/관리자', '해당 학생', '높음', '논문 제출', '최종 논문 제목이 승인되었습니다.'],
    ['TITLE_REJECTED', '최종 논문 제목 반려', '교수/관리자가 제목 반려', '교수/관리자', '해당 학생', '높음', '논문 제출', '최종 논문 제목이 반려되었습니다. 사유: [반려사유]'],

    # 3.4.1 심사위원 배정
    ['COMMITTEE_ASSIGNED', '심사위원 배정', '관리자가 심사위원 배정', '관리자', '해당 학생, 심사위원', '높음', '논문 심사', '[학생명]님의 논문 심사위원으로 배정되었습니다.'],
    ['COMMITTEE_CHANGED', '심사위원 변경', '관리자가 심사위원 변경', '관리자', '해당 학생, 기존/신규 심사위원', '높음', '논문 심사', '[기존 심사위원]에서 [신규 심사위원]으로 심사위원이 변경되었습니다.'],
    ['COMMITTEE_REMOVED', '심사위원 해제', '관리자가 심사위원 해제', '관리자', '해당 학생, 해제된 심사위원', '보통', '논문 심사', '심사위원 배정이 해제되었습니다.'],

    # 3.4.2 심사 일정
    ['SCHEDULE_REGISTERED', '심사 일정 등록', '관리자가 심사 일정 등록', '관리자', '해당 학생, 심사위원', '높음', '논문 심사', '논문 심사 일정이 등록되었습니다. 일시: [날짜] [시간], 장소: [장소]'],
    ['SCHEDULE_UPDATED', '심사 일정 변경', '관리자가 심사 일정 수정', '관리자', '해당 학생, 심사위원', '높음', '논문 심사', '논문 심사 일정이 변경되었습니다. 변경 후: [날짜] [시간]'],
    ['SCHEDULE_CANCELLED', '심사 일정 취소', '관리자가 심사 일정 취소', '관리자', '해당 학생, 심사위원', '높음', '논문 심사', '논문 심사 일정이 취소되었습니다.'],
    ['SCHEDULE_REMINDER_3DAYS', '심사 일정 알림 (3일 전)', '심사 3일 전 자동 발송', '시스템', '해당 학생, 심사위원', '높음', '논문 심사', '3일 후 논문 심사가 예정되어 있습니다.'],
    ['SCHEDULE_REMINDER_1DAY', '심사 일정 알림 (1일 전)', '심사 1일 전 자동 발송', '시스템', '해당 학생, 심사위원', '높음', '논문 심사', '내일 논문 심사가 예정되어 있습니다.'],
    ['SCHEDULE_REMINDER_2HOURS', '심사 일정 알림 (2시간 전)', '심사 2시간 전 자동 발송', '시스템', '해당 학생, 심사위원', '높음', '논문 심사', '2시간 후 논문 심사가 예정되어 있습니다.'],

    # 3.4.3 학위 논문 심사
    ['THESIS_REVIEW_STARTED', '학위 논문 심사 시작', '관리자가 심사 시작 처리', '관리자', '해당 학생, 심사위원', '높음', '논문 심사', '학위 논문 심사가 시작되었습니다.'],
    ['THESIS_REVIEW_COMPLETED', '학위 논문 심사 완료', '모든 심사위원이 심사 완료', '심사위원', '해당 학생, 관리자', '높음', '논문 심사', '학위 논문 심사가 완료되었습니다.'],
    ['THESIS_REVIEW_RESULT_PASS', '학위 논문 심사 통과', '심사 결과 통과 판정', '관리자', '해당 학생', '높음', '논문 심사', '학위 논문 심사 결과: 합격'],
    ['THESIS_REVIEW_RESULT_FAIL', '학위 논문 심사 불합격', '심사 결과 불합격 판정', '관리자', '해당 학생', '높음', '논문 심사', '학위 논문 심사 결과: 불합격'],
    ['THESIS_REVIEW_RESULT_CONDITIONAL', '학위 논문 심사 조건부 통과', '심사 결과 조건부 통과', '관리자', '해당 학생', '높음', '논문 심사', '학위 논문 심사 결과: 조건부 합격. 보완 사항: [보완사항]'],
    ['THESIS_REVIEW_COMMENT', '학위 논문 심사 의견 등록', '심사위원이 심사 의견 등록', '심사위원', '해당 학생, 관리자', '보통', '논문 심사', '심사위원이 심사 의견을 등록했습니다.'],

    # 3.4.4 학술지 논문 심사
    ['JOURNAL_REVIEW_STARTED', '학술지 논문 심사 시작', '관리자가 심사 시작 처리', '관리자', '해당 학생, 심사위원', '높음', '논문 심사', '학술지 논문 심사가 시작되었습니다.'],
    ['JOURNAL_REVIEW_COMPLETED', '학술지 논문 심사 완료', '모든 심사위원이 심사 완료', '심사위원', '해당 학생, 관리자', '높음', '논문 심사', '학술지 논문 심사가 완료되었습니다.'],
    ['JOURNAL_REVIEW_RESULT_PASS', '학술지 논문 심사 통과', '심사 결과 통과 판정', '관리자', '해당 학생', '높음', '논문 심사', '학술지 논문 심사 결과: 게재 승인'],
    ['JOURNAL_REVIEW_RESULT_FAIL', '학술지 논문 심사 불합격', '심사 결과 불합격 판정', '관리자', '해당 학생', '높음', '논문 심사', '학술지 논문 심사 결과: 게재 불가'],
    ['JOURNAL_REVIEW_COMMENT', '학술지 논문 심사 의견 등록', '심사위원이 심사 의견 등록', '심사위원', '해당 학생, 관리자', '보통', '논문 심사', '심사위원이 심사 의견을 등록했습니다.'],

    # 3.5 지도 교수 배정
    ['ADVISOR_ASSIGNED', '지도 교수 배정', '관리자가 지도 교수 배정', '관리자', '해당 학생, 교수', '높음', '지도 교수 배정', '[교수명] 교수님이 귀하의 지도 교수로 배정되었습니다.'],
    ['ADVISOR_CHANGED', '지도 교수 변경', '관리자가 지도 교수 변경', '관리자', '해당 학생, 기존/신규 교수', '높음', '지도 교수 배정', '지도 교수가 [기존 교수]에서 [신규 교수]로 변경되었습니다.'],
    ['ADVISOR_REMOVED', '지도 교수 해제', '관리자가 지도 교수 해제', '관리자', '해당 학생, 교수', '높음', '지도 교수 배정', '지도 교수 배정이 해제되었습니다.'],

    # 3.6 지도 단계
    ['STAGE_REGISTERED', '지도 단계 등록', '관리자/교수가 학생의 지도 단계 등록', '관리자/교수', '해당 학생', '높음', '지도 단계', '논문 지도 단계가 [단계명]으로 등록되었습니다.'],
    ['STAGE_UPDATED', '지도 단계 변경', '관리자/교수가 학생의 지도 단계 변경', '관리자/교수', '해당 학생', '높음', '지도 단계', '논문 지도 단계가 [이전 단계]에서 [새 단계]로 변경되었습니다.'],
    ['STAGE_COMPLETED', '지도 단계 완료', '관리자/교수가 단계 완료 처리', '관리자/교수', '해당 학생', '보통', '지도 단계', '[단계명] 단계가 완료되었습니다.'],

    # 3.7 시스템 관리
    ['WORK_SCHEDULE_DEADLINE', '업무 일정 마감 임박', '업무 마감 3일 전 자동 발송', '시스템', '관리자', '보통', '시스템 관리', '[업무명] 업무 마감이 3일 남았습니다.'],
    ['SUBMISSION_DEADLINE', '제출 마감 임박', '논문 제출 마감 7일/3일/1일 전', '시스템', '학생, 관리자', '높음', '시스템 관리', '논문 제출 마감이 7일 남았습니다.'],
    ['REVIEW_DEADLINE', '심사 마감 임박', '심사 마감 3일/1일 전', '시스템', '심사위원, 관리자', '높음', '시스템 관리', '[학생명] 논문 심사 마감이 3일 남았습니다.'],
    ['STATISTICS_REPORT', '통계 리포트 생성', '주간/월간 통계 리포트 생성', '시스템', '관리자', '낮음', '시스템 관리', '[기간] 통계 리포트가 생성되었습니다.'],
]

# DataFrame 생성
df_notifications = pd.DataFrame(notifications_data, columns=[
    '알림 코드', '알림 명칭', '발생 조건', '발신 주체', '수신 대상', '우선순위', '카테고리', '알림 메시지 예시'
])

# 2. 카테고리별 통계
category_stats = [
    ['논문 지도 계획', 6, '학기별 지도 계획 등록, 승인, 반려, 주차 추가/삭제'],
    ['논문 지도 활동', 11, '논문 지도 요청, 피드백, 실시간 예약'],
    ['논문 제출', 16, '논문 신청, 학위/학술지 논문 제출, 최종 제목 등록'],
    ['논문 심사', 21, '심사위원 배정, 심사 일정, 학위/학술지 논문 심사'],
    ['지도 교수 배정', 3, '지도 교수 배정, 변경, 해제'],
    ['지도 단계', 3, '지도 단계 등록, 변경, 완료'],
    ['시스템 관리', 4, '업무/제출/심사 마감 알림, 통계 리포트'],
]

df_category_stats = pd.DataFrame(category_stats, columns=['카테고리', '알림 개수', '주요 내용'])

# 3. 우선순위별 분류
priority_stats = [
    ['높음', 47, '즉시 확인 필요, 액션 필요', '빨간색 뱃지, 팝업 알림', '승인/반려, 제출 완료, 심사 결과'],
    ['보통', 14, '확인 필요하나 긴급하지 않음', '파란색 뱃지', '수정 알림, 의견 등록'],
    ['낮음', 3, '참고용 정보', '회색 뱃지', '통계 리포트, 주차 추가/삭제'],
]

df_priority = pd.DataFrame(priority_stats, columns=['우선순위', '개수', '설명', 'UI 표시', '예시'])

# 4. 발송 채널
channel_data = [
    ['인앱 알림', '모든 알림', '시스템 내 알림 센터', '읽음/안읽음 상태 관리', '필수'],
    ['이메일 알림', '우선순위 높음 알림', '10분마다 배치 발송', '사용자 설정 가능', 'Phase 2'],
    ['SMS 알림', '심사 일정 등 특정 알림', '실시간 발송', '사용자 설정 가능', 'Phase 3'],
    ['브라우저 푸시', '실시간 알림 필요 시', 'WebSocket 연결', '사용자 권한 필요', 'Phase 3'],
]

df_channels = pd.DataFrame(channel_data, columns=['채널', '대상 알림', '발송 방식', '비고', '구현 단계'])

# 5. 구현 우선순위
implementation_data = [
    ['Phase 1', '필수, 1차 구현', '인앱 알림 시스템, 논문 지도 요청/피드백, 논문 제출, 논문 심사, 학기별 지도 계획', '핵심 기능'],
    ['Phase 2', '2차 구현', '이메일 알림, 실시간 지도 예약, 심사 일정, 지도 교수 배정', '확장 기능'],
    ['Phase 3', '3차 구현', 'SMS/푸시 알림, 알림 설정 화면, 알림 통계 및 리포트', '선택 기능'],
]

df_implementation = pd.DataFrame(implementation_data, columns=['단계', '설명', '포함 기능', '비고'])

# Excel 파일 생성
wb = Workbook()
wb.remove(wb.active)

# 스타일 정의
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_sheet(ws, df, title):
    # 제목 추가
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center
    title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # 헤더 스타일
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = border

    # 데이터 추가
    for r_idx, row in enumerate(df.values, 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = border
            if c_idx == 1:
                cell.alignment = alignment_center
            else:
                cell.alignment = alignment_left

    # 열 너비 자동 조정
    from openpyxl.utils import get_column_letter
    for col_idx, column_title in enumerate(df.columns, 1):
        max_length = len(str(column_title))
        for r_idx in range(3, len(df) + 3):
            cell_value = ws.cell(row=r_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

# 1. 알림 전체 목록 시트
ws1 = wb.create_sheet('알림 전체 목록')
style_sheet(ws1, df_notifications, '알림 시스템 - 전체 알림 목록 (총 64개)')

# 2. 카테고리별 통계 시트
ws2 = wb.create_sheet('카테고리별 통계')
style_sheet(ws2, df_category_stats, '카테고리별 알림 통계')

# 3. 우선순위별 분류 시트
ws3 = wb.create_sheet('우선순위별 분류')
style_sheet(ws3, df_priority, '우선순위별 알림 분류')

# 4. 발송 채널 시트
ws4 = wb.create_sheet('발송 채널')
style_sheet(ws4, df_channels, '알림 발송 채널')

# 5. 구현 우선순위 시트
ws5 = wb.create_sheet('구현 우선순위')
style_sheet(ws5, df_implementation, '알림 시스템 구현 우선순위')

# 6. 요약 시트
ws_summary = wb.create_sheet('요약', 0)
summary_data = [
    ['항목', '내용'],
    ['총 알림 개수', '64개'],
    ['알림 카테고리', '7개 (논문 지도 계획, 논문 지도 활동, 논문 제출, 논문 심사, 지도 교수 배정, 지도 단계, 시스템 관리)'],
    ['우선순위', '3단계 (높음: 47개, 보통: 14개, 낮음: 3개)'],
    ['발송 채널', '4종류 (인앱, 이메일, SMS, 브라우저 푸시)'],
    ['구현 단계', 'Phase 1 (필수), Phase 2 (확장), Phase 3 (선택)'],
    ['문서 작성일', '2026-01-20'],
    ['버전', '1.0'],
]

for r_idx, (key, value) in enumerate(summary_data, 1):
    cell_key = ws_summary.cell(row=r_idx, column=1)
    cell_value = ws_summary.cell(row=r_idx, column=2)
    cell_key.value = key
    cell_value.value = value

    if r_idx == 1:
        cell_key.font = header_font
        cell_value.font = header_font
        cell_key.fill = header_fill
        cell_value.fill = header_fill
    else:
        cell_key.font = Font(bold=True)

    cell_key.border = border
    cell_value.border = border
    cell_key.alignment = alignment_center
    cell_value.alignment = alignment_left

ws_summary.column_dimensions['A'].width = 20
ws_summary.column_dimensions['B'].width = 100

# 파일 저장
wb.save('notification-system-requirements.xlsx')
print('Excel 파일이 성공적으로 생성되었습니다: notification-system-requirements.xlsx')
