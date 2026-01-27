import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 엑셀 워크북 생성
wb = openpyxl.Workbook()
wb.remove(wb.active)  # 기본 시트 삭제

# 스타일 정의
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 수용 여부별 색상
accept_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # 녹색
partial_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # 노란색
reject_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # 주황색

def create_sheet(wb, sheet_name, data):
    """시트 생성 및 데이터 입력"""
    ws = wb.create_sheet(title=sheet_name)

    # 헤더 설정
    headers = ["번호", "화면명", "기능명", "논의 내용", "구현 방안", "수용 여부", "확인필요사항", "Phase"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 데이터 입력
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border

            # 수용 여부에 따라 색상 적용
            if col_num == 6:  # 수용 여부 컬럼
                if "수용" in str(value) and "불가" not in str(value) and "부분" not in str(value):
                    cell.fill = accept_fill
                elif "부분" in str(value) or "후순위" in str(value):
                    cell.fill = partial_fill
                elif "불가" in str(value):
                    cell.fill = reject_fill

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 12

    # 첫 행 고정
    ws.freeze_panes = 'A2'

# 관리자 화면 데이터
admin_data = [
    [1, "관리자", "단계지도유형관리", "본부 공통관리 3개 항목(논문작성계획서, 예비심사, 본심사) 이외 관리 이슈 없음. 화면 삭제 제안", "지도단계유형관리 화면 삭제, 심사유형만 존재", "부분 수용 - 확인 필요", "화면 필요 여부 한양대 재확인 필요", "Phase 1"],
    [2, "관리자", "심사평가기준 등록", "점수형, P/F형 이외 서술형, 척도형 평가 추가 요청", "서술형, 척도형 평가 추가. 종합평정 책임자 심사 유형별 자동 구분", "수용", "", "Phase 1"],
    [3, "관리자", "지도단계등록 - 우산 개념", "큰 카테고리 내 세부 단계 등록. 카테고리별 신청 1회, 세부 차수별 제출/심사 반복", "단계 카테고리와 세부 단계 분리 관리", "수용", "", "Phase 1"],
    [4, "관리자", "지도단계등록 - 심사 유형", "단계별 지도교수 단독 심사 vs 심사위원회 심사 정의 필요", "지도단계 정보에 심사 유형 항목 추가. 중간 단계는 지도교수 단독, 최종 단계는 심사위원회", "수용", "제3의 케이스 존재 여부 확인", "Phase 1"],
    [5, "관리자", "지도단계별 일정 등록", "일정 유형에 따라 일정 등록", "기존 기능 유지", "수용", "", "Phase 1"],
    [6, "관리자", "지도교수-학생 배정/배치", "기간계 시스템에서 배정, 논문지도시스템에서 일 배치로 인입", "기존 기능 유지", "수용", "", "Phase 1"],
    [7, "관리자", "학기 개념 추가", "대상자 관리를 위한 학기 구분 필요. 지도교수 배정은 연속성 유지, 단계별은 학기 구분", "학생-단계 매칭 정보에 학기 항목 추가. 졸업 미달 학생 자동 연장", "수용", "", "Phase 2"],
    [8, "관리자", "심사위원 배정", "지도교수 단독 케이스 심사위원장 옵션화 제안 → 수용 불가", "심사 유형 분리. 지도교수 단독 심사 시 심사위원 배정 화면에서 제외", "수용 불가 (대안 수용)", "", "Phase 1"],
    [9, "관리자", "심사위원 재사용", "예비심사 N차, 본심사 N차는 기존 위원 동일 진행. 기존 구성 불러오기 기능 필요", "이전 차수 구성 내역 불러오기 기능 추가. 일부 위원 변경 가능", "수용", "", "Phase 1"],
    [10, "관리자", "관리자 화면 개선", "단계별 신청/제출/심사 내역 구분 관리. 대상자, 제출 여부, 성적 입력, 합격 여부 통합 조회", "단계별 통합 조회 화면, 필터링 기능 구현", "수용", "", "Phase 2"],
]

# 학생 화면 데이터
student_data = [
    [1, "학생", "심사 신청", "각 단계별 신청-제출-심사 반복 구조 유지", "현재 구조 유지", "수용", "", "Phase 1"],
    [2, "학생", "논문 제출 - 기본", "교수에게 승인 요청 산출물 등록", "기존 기능 유지", "수용", "", "Phase 1"],
    [3, "학생", "논문 제출 - 부속서류", "N차 심사, 재심 시 수정전후대비표 등 부속서류 제출", "논문+부속서류 동시 업로드 기능 추가", "수용", "", "Phase 1"],
    [4, "학생", "재심 - 재제출", "재제출 시 UI 명확화, 부속서류 업로드", "학생 제출 화면 재제출 여부 명확 표시", "수용", "", "Phase 1"],
    [5, "학생", "학기별 논문지도계획 작성", "학기별 지도계획 작성 및 승인 신청", "기존 기능 유지", "수용", "", "Phase 1"],
    [6, "학생", "실시간지도 신청", "교수가 설정한 일정 슬롯 중 선택하여 예약 신청", "기존 기능 유지", "수용", "", "Phase 1"],
    [7, "학생", "논문지도대상 파일 등록", "피드백 대상 파일 게시. 지도단계 기본값 설정", "기존 기능 유지", "수용", "", "Phase 1"],
    [8, "학생", "파일 수정 제한", "교수 피드백 전 파일 수정 가능. 열람건수 0일 때만 수정", "열람건수 기반 수정 가능 여부 판단", "수용", "", "Phase 1"],
    [9, "학생", "파일 업로드 - 피드백 반영", "교수 피드백 확인 후 수정 파일 재업로드", "기존 기능 유지", "수용", "", "Phase 1"],
    [10, "학생", "영역지정 피드백 기능", "학생도 PDF 영역지정하여 피드백 업로드 가능", "학생용 영역지정 피드백 기능 추가, 양방향 소통 UI", "수용", "", "Phase 3"],
    [11, "학생", "파일 버전 관리", "이전 버전 피드백과 새 문서 피드백 매핑 방법 없음", "Phase 1: 버전 번호+이전 참조 / Phase 3: 상세 피드백 매핑", "부분 수용 - Phase별 구분", "", "Phase 1/3"],
    [12, "학생", "플로우 시각화", "대시보드에 플로우 이미지 표시, 현재 진행 단계 시각화", "학생 대시보드 플로우 시각화 추가", "수용", "", "Phase 2"],
    [13, "학생", "메뉴 구조 개선", "단계 유형(논문작성계획서, 예비심사, 본심사)으로 구분, 세부 단계 하위 표시", "학생 제출 메뉴 구조 개선", "수용", "", "Phase 1"],
]

# 교수 화면 데이터
professor_data = [
    [1, "교수", "학기별 논문지도계획 승인", "지도계획 확인/수정 및 승인. 승인 후 수정 불가", "기존 기능 유지", "수용", "", "Phase 1"],
    [2, "교수", "실시간지도 일정 등록", "미팅 가능 일정 슬롯 등록(30분~120분)", "기존 기능 유지", "수용", "", "Phase 1"],
    [3, "교수", "실시간지도 승인/거절", "학생 예약 승인/거절. 거절 시 사유 입력, 슬롯 복구", "기존 기능 유지", "수용", "", "Phase 1"],
    [4, "교수", "거절 이력 관리", "거절 항목 별도 저장하여 조회 가능", "거절 이력 저장 기능 추가", "수용", "", "Phase 1"],
    [5, "교수", "줌 미팅 생성", "온라인 실시간지도 승인 시 줌 미팅 생성 및 주소 반환", "기존 기능 유지", "수용", "", "Phase 1"],
    [6, "교수", "자동 완료처리", "완료된 일정 자동 완료처리", "자동 완료처리 적용", "수용", "", "Phase 1"],
    [7, "교수", "캘린더 UI", "캘린더 UI 개선", "후순위 진행", "후순위", "", "Phase 3"],
    [8, "교수", "논문 심사 - 조건부합격", "합격, 조건부합격, 불합격 3단계 구분", "조건부합격 시 재심사 프로세스 진행", "수용", "", "Phase 1"],
    [9, "교수", "재심 방식 - 대면", "대면 일정 지정, 심사위원회 전체 심의", "재심(대면) 프로세스 구현", "수용", "", "Phase 1"],
    [10, "교수", "재심 방식 - 서면", "심사위원장 승인 또는 지도교수 위임 승인", "Phase 2 이후 검토", "부분 수용 - 서면평가 제외", "", "Phase 2"],
    [11, "교수", "논문지도단계 승급", "모든 단계 심사-승인 프로세스", "지도교수 1인: 승인 화면 / 심사위원회: 평가표 화면", "수용 (구현 방식 변경)", "", "Phase 1"],
    [12, "교수", "교수 피드백", "PDF 파일 확인, 판서 및 영역 지정 피드백", "기존 기능 유지", "수용", "", "Phase 1"],
    [13, "교수", "단계 정보 저장", "피드백 작성 시점 단계 정보 저장하여 히스토리 보존", "피드백 정보에 작성 시점 단계 항목 추가", "수용", "", "Phase 1"],
    [14, "교수", "플로우 시각화", "학생별 플로우 시각화. 학생 정보 클릭 시 팝업 표시", "학생 목록 현재 단계 표시, 팝업 플로우 시각화", "수용", "", "Phase 2"],
    [15, "교수", "메뉴 구조 개선", "제출 내역 조회 vs 심사 평가 메뉴 분리", "제출 내역 조회: 모든 산출물 / 심사 평가: 평가 필요 단계만", "수용", "", "Phase 1"],
]

# 시트 생성
create_sheet(wb, "관리자 화면", admin_data)
create_sheet(wb, "학생 화면", student_data)
create_sheet(wb, "교수 화면", professor_data)

# 요약 시트 생성
summary_ws = wb.create_sheet(title="요약", index=0)

# 요약 헤더
summary_headers = ["구분", "수용", "부분 수용", "수용 불가", "후순위", "합계"]
for col_num, header in enumerate(summary_headers, 1):
    cell = summary_ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# 요약 데이터
summary_data = [
    ["관리자 화면", 8, 1, 1, 0, 10],
    ["학생 화면", 10, 1, 0, 0, 11],  # 파일 버전 관리는 부분 수용
    ["교수 화면", 12, 1, 0, 1, 14],  # 캘린더 UI는 후순위
    ["합계", 30, 3, 1, 1, 35],
]

for row_num, row_data in enumerate(summary_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = summary_ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

        if row_num == 5:  # 합계 행
            cell.font = Font(bold=True)
            cell.fill = subheader_fill

# 요약 시트 컬럼 너비 조정
for col in range(1, 7):
    summary_ws.column_dimensions[get_column_letter(col)].width = 15

# Phase별 요약 추가
summary_ws.cell(row=7, column=1).value = "Phase별 요약"
summary_ws.cell(row=7, column=1).font = Font(bold=True, size=12)

phase_headers = ["Phase", "항목 수", "예상 공수(일)"]
for col_num, header in enumerate(phase_headers, 1):
    cell = summary_ws.cell(row=8, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

phase_data = [
    ["Phase 1 (필수)", 25, 21],
    ["Phase 2 (개선)", 5, 16],
    ["Phase 3 (UX)", 5, 11],
    ["합계", 35, 48],
]

for row_num, row_data in enumerate(phase_data, 9):
    for col_num, value in enumerate(row_data, 1):
        cell = summary_ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

        if row_num == 12:  # 합계 행
            cell.font = Font(bold=True)
            cell.fill = subheader_fill

# 파일 저장
wb.save("요구사항_수용여부_정리.xlsx")
print("엑셀 파일이 생성되었습니다: 요구사항_수용여부_정리.xlsx")
